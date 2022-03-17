# coding = utf-8
"""
Author: DaisyZhou

date: 2019/11/27 18:14

单因子测试结果的输出
"""
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import seaborn as sns
import xlwings as xw
from openpyxl.utils import get_column_letter, column_index_from_string
from xlwings import constants

# from . import date_utils #封装

import factor_performance #封装时注释
import date_utils #封装时注释

def formatrange(ws, xlrange, numformt, align, **kwargs):
    # http: // autoofficepy.blogspot.com / 2017 / 12 / xlwings - formatting.html

    if numformt != '': #数字格式
        ws.range(xlrange).api.NumberFormat = numformt

    ### set up alignment
    if align =='Center':
        ws.range(xlrange).api.HorizontalAlignment = constants.HAlign.xlHAlignCenter
    if align =='Right':
        ws.range(xlrange).api.HorizontalAlignment = constants.HAlign.xlHAlignRight
    if align =='Left':
        ws.range(xlrange).api.HorizontalAlignment = constants.HAlign.xlHAlignLeft

    if 'if_Bold' in kwargs.keys():
        ws.range(xlrange).api.Font.Bold = kwargs['if_Bold']

def print_table(table, name=None, fmt=None):
    """
    Pretty print a pandas DataFrame.

    Uses HTML output if running inside Jupyter Notebook, otherwise
    formatted text output.

    Parameters
    ----------
    table : pd.Series or pd.DataFrame
        Table to pretty-print.
    name : str, optional
        Table name to display in upper left corner.
    fmt : str, optional
        Formatter to use for displaying table elements.
        E.g. '{0:.2f}%' for displaying 100 as '100.00%'.
        Restores original setting after displaying.
    """
    if isinstance(table, pd.Series):
        table = pd.DataFrame(table)

    if isinstance(table, pd.DataFrame):
        table.columns.name = name

    prev_option = pd.get_option('display.float_format')
    if fmt is not None:
        pd.set_option('display.float_format', lambda x: fmt.format(x))

    display(table)

    if fmt is not None:
        pd.set_option('display.float_format', prev_option)
    return table

def plot_ic_all(ic_inf, title, dataname):
    fig1 = plt.figure(constrained_layout=True)
    gs = fig1.add_gridspec(ncols=2, nrows=2)
    fig1_ax1 = fig1.add_subplot(gs[0, 0])
    #子图1：IC分析结果
    # fig1_ax1 = plt.subplot2grid(gridsize, (0, 0))
    temp = [[title + '_mean', title + '_std', title + 'IR', title + '_t', 'pct_pos', 'pct_neg'], ic_inf[dataname + '_result'].to_list()]
    fig1_ax1.axis('off')
    table = fig1_ax1.table(cellText=list(zip([title + '_mean', title + '_std', title + 'IR', title + '_t', 'pct_pos', 'pct_neg'],
                                list(map(lambda x: round(x,4), ic_inf[dataname + '_result'].to_list())))), loc = 'center', fontsize=16,
                      rowLoc='center')
    table.scale(1, 1.5)
    fig1_ax1.set_title(title + ' result')


    fig1_ax2 = fig1.add_subplot(gs[0, 1])
    sns.distplot(ic_inf['ic'].replace(np.nan, 0.), norm_hist=True)
    fig1_ax2.text(0.05, 0.8, "Mean %.3f \n Std. %.3f" % (ic_inf[dataname].mean(), ic_inf[dataname].std()),
           fontsize=6,
           bbox={'facecolor': 'white', 'alpha': 1, 'pad': 5},
           transform=fig1_ax2.transAxes,
           verticalalignment='top')
    fig1_ax2.axvline(ic_inf[dataname].mean(), color='w', linestyle='dashed', linewidth=2)
    fig1_ax2.set_title(title + ' hist')


    fig1_ax3 = fig1.add_subplot(gs[1, :])
    date_utils.chg_idx2datestr(ic_inf[dataname]).plot(alpha=0.7, ax=fig1_ax3, lw=0.7, color='steelblue')
    date_utils.chg_idx2datestr(ic_inf[dataname].rolling(window=3).mean()).plot(alpha=0.8, ax=fig1_ax3, lw=2, color='forestgreen')
    fig1_ax3.set(ylabel=title, xlabel="")
    fig1_ax3.set_title(
        title)
    fig1_ax3.axhline(0.0, linestyle='-', color='black', lw=1, alpha=0.8)
    fig1_ax3.legend([title, '3 periods moving avg'], loc='upper right', prop={'size': 6})


    return fig1

def dict2string(dic):
    #将字典内容转换为字符串
    s1 = dic.items()
    lst = []
    # k = []
    # v = []
    for key, value in s1:
        s3 = "%s=%s" % (key, value)
        # k_i = "%s" % (key)
        # v_i = "%s" % (value)
        # k.append(k_i)
        # v.append(v_i)
        lst.append(s3)
    return lst #, k, v

def get_rs_cols(ws):
    global r_count, c_count
    r_count = ws.api.UsedRange.Rows.count
    c_count = ws.api.UsedRange.Columns.count

def save_config(result_factor_analyse):
    global wb, current_ws
    ws = wb.sheets.add('配置信息')
    factor_setting = result_factor_analyse['factor_setting']
    prepro_setting = result_factor_analyse['prepro_setting']
    execute_setting = result_factor_analyse['execute_setting']

    ws.range('A1').value = '因子信息'
    ws.range('A1').api.Font.Bold = True
    ws.range('A2').value = pd.Series(factor_setting)

    ws.range('A5').value = '预处理方式'
    ws.range('A5').api.Font.Bold = True
    ws.range('A6').options(transpose=True).value = dict2string(prepro_setting) #纵向

    get_rs_cols(ws)
    ws.range('A{}'.format(r_count+2)).value = '运行信息'
    ws.range('A{}'.format(r_count+2)).api.Font.Bold = True
    ws.range('A{}'.format(r_count+3)).options(transpose=True).value = dict2string(execute_setting)  # 纵向

    current_ws = '配置信息'

def save_ic(result_analyse_effect, title=''):
    global wb, current_ws
    # result_analyse_effect = result_factor_analyse['analyse_effect']
    ws = wb.sheets.add(title +'IC分析', after=current_ws)
    ws.range('A1').value = 'IC'
    formatrange(ws, 'A1', "", 'Center', if_Bold=True)
    ws.range('B1').value = result_analyse_effect['ic_inf']['ic_result']
    #设置数字格式
    formatrange(ws, 'C1:C4', "0.0000", 'Center')
    formatrange(ws, 'C5:C6', "0.00%", 'Center')

    ws.range('E1').value = 'IC序列'
    ws.range('F1:G1').value = ['调仓日', 'IC值']
    ws.range('F2').value = date_utils.chg_idx2datestr(result_analyse_effect['ic_inf']['ic'])
    #
    formatrange(ws, 'E1', "", 'Center', if_Bold=True)
    formatrange(ws, 'F:F', "", 'Center')
    formatrange(ws, 'G:G', "0.0000", 'Center')


    #RankIC
    ws.range('I1').value = 'rankIC'
    ws.range('J1').value = result_analyse_effect['ic_inf']['rank_ic_result']
    ws.range('M1').value = 'rankIC序列'
    ws.range('N1:O1').value = ['调仓日', 'rankIC值']
    ws.range('N2').value = date_utils.chg_idx2datestr(result_analyse_effect['ic_inf']['rank_ic'])

    formatrange(ws, 'I1', "", 'Center', if_Bold=True)
    formatrange(ws, 'K1:K4', "0.0000", 'Center')
    formatrange(ws, 'K5:K6', "0.00%", 'Center')
    #
    formatrange(ws, 'M1', "", 'Center', if_Bold=True)
    formatrange(ws, 'N:N', "", 'Center')
    formatrange(ws, 'O:O', "0.0000", 'Center')

    #因子自相关性
    ws.range('Q1').value = '因子秩自相关性'
    ws.range('R1:S1').value = ['调仓日', 'factor_rank_autocorr']
    ws.range('R2').value = date_utils.chg_idx2datestr(result_analyse_effect['ic_inf']['factor_rank_autocorr'])
    formatrange(ws, 'Q1', "", 'Center', if_Bold=True)
    formatrange(ws, 'R:R', "", 'Center')
    formatrange(ws, 'S:S', "0.0000", 'Center')

    #画图：
    sns.set()
    fig1 = plot_ic_all(result_analyse_effect['ic_inf'], 'IC', 'ic')
    ws.pictures.add(fig1, name='ICPlot', update=True, left=ws.range('A13').left, top=ws.range('A13').top)
    plt.close(fig1)
    fig2 = plot_ic_all(result_analyse_effect['ic_inf'], 'Rank_IC', 'rank_ic')
    ws.pictures.add(fig2, name='Rank_ICPlot', update=True, left=ws.range('L13').left, top=ws.range('L13').top)
    plt.close(fig2)
    f, a = plt.subplots(1, 1, figsize=(8, 3))
    date_utils.chg_idx2datestr(result_analyse_effect['ic_inf']['factor_rank_autocorr']).plot(alpha=1, ax=a, lw=1, color='steelblue')
    a.set_title('factor_rank_autocorr')
    ws.pictures.add(f, name='factor_rank_autocorr', update=True, left=ws.range('X13').left, top=ws.range('X13').top)
    plt.close(f)

    current_ws = title + 'IC分析'

def save_fac_ret(result_analyse_effect, title=''):
    global wb, current_ws
    result_fac_ret = result_analyse_effect['fac_return_inf']
    ws = wb.sheets.add(title + '因子收益', after=current_ws)
    ws.range('A1').value = '因子收益'
    ws.range('B1').value = result_fac_ret['fac_return_result']

    formatrange(ws, 'A1', '', 'Center', if_Bold=True)
    formatrange(ws, 'C1:C4', "0.0000", 'Center')
    formatrange(ws, 'C5:C6', "0.00%", 'Center')

    ws.range('E1').value = '因子收益序列'
    ws.range('F1').value = date_utils.chg_idx2datestr(result_fac_ret['fac_return_value'])
    formatrange(ws, 'E1', '', 'Center', if_Bold=True)
    formatrange(ws, 'G:H', "0.00", 'Center')
    formatrange(ws, 'I:I', "0.00%", 'Center')


    current_ws = title + '因子收益'
    f, a = plt.subplots(1, 1, figsize=(8, 3))
    date_utils.chg_idx2datestr(result_analyse_effect['fac_return_inf']['fac_return_value'].loc[:, 'fac_ret']).plot(alpha=1, ax=a, lw=1, color='steelblue')
    a.set_title('fac_return_value')
    a.axhline(0.0, linestyle='-', color='black', lw=1, alpha=0.8)
    ws.pictures.add(f, name='fac_return_value', update=True, left=ws.range('K1').left, top=ws.range('K1').top)
    plt.close(f)

def save_longshort_inf(result_analyse_effect, title=''):
    global wb, current_ws
    ws = wb.sheets.add(title + '多空收益', after=current_ws)
    get_rs_cols(ws)
    longshort_inf = result_analyse_effect['longshort_inf']
    ws.range(get_column_letter(c_count + 0) + '1').value = '多空表现_总体'
    ws.range(get_column_letter(c_count + 1) + '1').value = longshort_inf['eva_l_s_ls']
    #
    formatrange(ws, 'A1', "", 'Center', if_Bold=True)
    formatrange(ws, 'C:E', "", 'Center')
    formatrange(ws, 'C4:E4', "0.00", 'Center')
    formatrange(ws, 'C7:E8', "0.00", 'Center')
    formatrange(ws, 'C1:E8', "0.00%", 'Center')

    get_rs_cols(ws)
    ws.range(get_column_letter(c_count + 2) + '1').value = '多空表现_分年'
    formatrange(ws, get_column_letter(c_count + 2) + '1', "", 'Center', if_Bold=True)
    # ws.range(get_column_letter(c_count + 3) + '1').value = '多头超额'
    # ws.range(get_column_letter(c_count + 4) + '1').value = longshort_inf['eva_yearly_l_s_ls']['多头超额'].set_index('Year')
    # ws.range(get_column_letter(c_count + 3) + '1').value = '空头超额'


    ws.range(get_column_letter(c_count + 3) + '1').value = pd.concat(
        [pd.DataFrame(list(['']), columns=[''], index=['多头超额']),
         longshort_inf['eva_yearly_l_s_ls']['多头超额'].set_index('Year'),
         pd.DataFrame(list(['']), columns=[''], index=['空头超额']),
         longshort_inf['eva_yearly_l_s_ls']['空头超额'].set_index('Year'),
         pd.DataFrame(list(['']), columns=[''], index=['多空']),
         longshort_inf['eva_yearly_l_s_ls']['多空'].set_index('Year')], axis=0, sort=False)

    #
    formatrange(ws, get_column_letter(c_count + 4) + ':' + get_column_letter(c_count + 11),
                "0.00%", 'Center')
    formatrange(ws, get_column_letter(c_count + 7) + ':' + get_column_letter(c_count + 7),
                "0.00", 'Center')
    formatrange(ws, get_column_letter(c_count + 9) + ':' + get_column_letter(c_count + 11),
                "0.00", 'Center')
    formatrange(ws, get_column_letter(c_count + 11) + ':' + get_column_letter(c_count + 13),
                "", 'Center')


    # 各期收益
    get_rs_cols(ws)
    ws.range(get_column_letter(c_count + 2) + '1').value = '每期收益'
    ws.range(get_column_letter(c_count + 3) + '1').value = date_utils.chg_idx2datestr(
        longshort_inf['period_ret_l_s_ls'])
    #
    formatrange(ws, get_column_letter(c_count + 3) + ':' + get_column_letter(c_count + 3),
                "", 'Center')
    formatrange(ws, get_column_letter(c_count + 4) + ':' + get_column_letter(c_count + 6),
                "0.00%", 'Center')

    # 净值
    get_rs_cols(ws)
    ws.range(get_column_letter(c_count + 2) + '1').value = '净值'
    ws.range(get_column_letter(c_count + 3) + '1').value = date_utils.chg_idx2datestr(longshort_inf['net_l_s_ls'])
    current_ws = title + '多空收益'
    formatrange(ws, get_column_letter(c_count + 4) + ':' + get_column_letter(c_count + 8),
                "0.00", 'Center')

    #画图：
    #1.每期多空收益
    sns.set()
    f, a = plt.subplots(1, 1, figsize=(8, 3))
    plt.bar(np.arange(len(longshort_inf['period_ret_l_s_ls'])), date_utils.chg_idx2datestr(longshort_inf['period_ret_l_s_ls']['多空']))
    a.axhline(0.0, linestyle='-', color='black', lw=1, alpha=0.8)
    a.set_title('long_short period return')
    ws.pictures.add(f, name='long_short period return', update=True, left=ws.range('A12').left, top=ws.range('A12').top)
    plt.close(f)

    #2. 净值（多头超额、空头超额、多空）
    sns.set()
    f, a = plt.subplots(1, 1, figsize=(8, 3))
    date_utils.chg_idx2datestr(longshort_inf['net_l_s_ls']['多头超额']).plot(alpha=1, ax=a, lw=1, label='long excess return', color='lightcoral')
    date_utils.chg_idx2datestr(longshort_inf['net_l_s_ls']['空头超额']).plot(alpha=1, ax=a, lw=1, label='short excess return', color='mediumaquamarine')
    date_utils.chg_idx2datestr(longshort_inf['net_l_s_ls']['多空']).plot(alpha=1, ax=a, lw=1, label='long short excess return', color='cornflowerblue')
    plt.legend(loc='lower left', fontsize='x-small', bbox_to_anchor=(1, 0.5))
    a.set_title('long short net')
    plt.close(f)
    ws.pictures.add(f, name='long short net', update=True, left=ws.range('A34').left, top=ws.range('A34').top)

def save_group_inf(result_analyse_effect, title=''):
    global wb, current_ws
    ws = wb.sheets.add(title + '分组收益', after=current_ws)
    # result_analyse_effect = result_factor_analyse['analyse_effect']
    group_ret_inf = result_analyse_effect['group_ret_inf']
    # 分组超额评价
    ws.range('A1').value = '分组超额收益'
    ws.range('B1').value = group_ret_inf['group_eva_exc']
    ws.range('C1').value = ['组'+str(i) for i in group_ret_inf['group_eva_exc'].columns.to_list()]
    get_rs_cols(ws)
    #
    formatrange(ws, 'A1', "", 'Center', if_Bold=True)
    formatrange(ws, 'C2' + ':' + get_column_letter(c_count) + '8', "0.00%", 'Center')
    formatrange(ws, 'C4' + ':' + get_column_letter(c_count) + '4', "0.00", 'Center')
    formatrange(ws, 'C6' + ':' + get_column_letter(c_count) + '8', "0.00", 'Center')
    formatrange(ws, 'C' + ':' + get_column_letter(c_count), "", 'Center')

    ws.range('A{}'.format(r_count+2)).value = '分组绝对收益'
    ws.range('B{}'.format(r_count+2)).value = group_ret_inf['group_eva_abs']
    ws.range('C{}'.format(r_count+2)).value = ['组'+str(i) for i in group_ret_inf['group_eva_abs'].columns.to_list()]
    formatrange(ws, 'A{}'.format(r_count+2), "", 'Center', if_Bold=True)
    formatrange(ws, 'C{}'.format(r_count+2) + ':' +
                get_column_letter(c_count) + str(8 +r_count+1), "0.00%", 'Center')
    formatrange(ws, 'C' + str(4+r_count+1) + ':' + get_column_letter(c_count) + str(4+r_count+1), "0.00", 'Center')
    formatrange(ws, 'C' + str(6+r_count+1) + ':' + get_column_letter(c_count) + str(8+r_count+1), "0.00", 'Center')


    get_rs_cols(ws)
    ws.range('B{}'.format(r_count + 2)).value = '换手率'
    ws.range('C{}'.format(r_count + 2)).value = list(group_ret_inf['turn_fac'].mean())
    formatrange(ws, 'C{}'.format(r_count + 2) + ':' + get_column_letter(c_count) + str(r_count + 2),
                "0.00%", 'Center')

    #分组净值-不对冲
    get_rs_cols(ws)
    ws.range(get_column_letter(c_count + 2) + '1').value = '分组不对冲净值'
    ws.range(get_column_letter(c_count + 3) + '1').value = date_utils.chg_idx2datestr(pd.merge(group_ret_inf['daily_net_simp'], result_analyse_effect['longshort_inf']['net_l_s_ls']['多空'], left_index = True, right_index = True))
    ws.range(get_column_letter(c_count + 4) + '1').value = ['组'+str(i) for i in group_ret_inf['daily_net_simp'].columns.to_list()] + ['多空']

    get_rs_cols(ws)
    #g
    g_num = len(group_ret_inf['daily_net_simp'].columns)
    formatrange(ws, get_column_letter(c_count - g_num) + ':' + get_column_letter(c_count), '0.00', 'Center')

    # 分组净值-对冲
    ws.range(get_column_letter(c_count + 2) + '1').value = '分组对冲净值'
    #和多空净值合并
    ws.range(get_column_letter(c_count + 3) + '1').value = date_utils.chg_idx2datestr(pd.merge(group_ret_inf['daily_excnet_simp'], result_analyse_effect['longshort_inf']['net_l_s_ls']['多空'], left_index = True, right_index = True))
    ws.range(get_column_letter(c_count + 4) + '1').value = ['组' + str(i) for i in
                                                            group_ret_inf['daily_excnet_simp'].columns.to_list()] + ['多空']
    get_rs_cols(ws)
    formatrange(ws, get_column_letter(c_count - g_num) + ':' + get_column_letter(c_count), '0.00', 'Center')

    # 分组各期收益
    ws.range(get_column_letter(c_count + 2) + '1').value = '分组各期收益(不对冲)'
    ws.range(get_column_letter(c_count + 3) + '1').value = date_utils.chg_idx2datestr(pd.merge(group_ret_inf['group_ret'], result_analyse_effect['longshort_inf']['longshort_ret'].to_frame(), left_index = True, right_index = True))
    ws.range(get_column_letter(c_count + 4) + '1').value = ['组' + str(i) for i in
                                                            group_ret_inf['group_ret'].columns.to_list()] + ['多空']
    get_rs_cols(ws)
    formatrange(ws, get_column_letter(c_count - g_num) + ':' + get_column_letter(c_count), '0.00%', 'Center')

    # 各期每组股票数目
    ws.range(get_column_letter(c_count + 2) + '1').value = '各期每组股票数目'
    ws.range(get_column_letter(c_count + 3) + '1').value = date_utils.chg_idx2datestr(group_ret_inf['group_num'])
    ws.range(get_column_letter(c_count + 4) + '1').value = ['组' + str(i) for i in
                                                            group_ret_inf['group_num'].columns.to_list()]
    get_rs_cols(ws)
    formatrange(ws, get_column_letter(c_count - g_num + 1) + ':' + get_column_letter(c_count), '0', 'Center')

    # 各期分组排名
    ws.range(get_column_letter(c_count + 2) + '1').value = '各期分组排名'
    ws.range(get_column_letter(c_count + 3) + '1').value = date_utils.chg_idx2datestr(group_ret_inf['group_rank'])
    ws.range(get_column_letter(c_count + 4) + '1').value = ['组' + str(i) for i in
                                                            group_ret_inf['group_rank'].columns.to_list()]
    get_rs_cols(ws)
    formatrange(ws, get_column_letter(c_count - g_num + 1) + ':' + get_column_letter(c_count), '0', 'Center')
    current_ws = title + '分组收益'

    #画图：
    #1. 分组年化收益
    sns.set()
    f, a = plt.subplots(1, 1, figsize=(5, 2))
    plt.bar(np.arange(group_ret_inf['group_eva_exc'].shape[1]), group_ret_inf['group_eva_exc'].loc['AnnualRt',:],
            tick_label = group_ret_inf['group_eva_exc'].columns.to_list())
    a.set_title('Group Return')
    plt.close(f)
    ws.pictures.add(f, name='group returnt', update=True, left=ws.range('A25').left, top=ws.range('A25').top)

    #2. 换手率
    f, a = plt.subplots(1, 1, figsize=(5, 2))
    plt.bar(np.arange(group_ret_inf['group_eva_exc'].shape[1]), group_ret_inf['turn_fac'].mean(),
            tick_label=group_ret_inf['group_eva_exc'].columns.to_list())
    a.set_title('Group Turnover')
    plt.close(f)
    ws.pictures.add(f, name='group turnover', update=True, left=ws.range('A40').left, top=ws.range('A40').top)

    #3. 分组净值
    sns.set()
    sns.set_palette(sns.cubehelix_palette(group_ret_inf['group_eva_exc'].shape[1]))
    f, a = plt.subplots(1, 1, figsize=(9, 3))
    date_utils.chg_idx2datestr(group_ret_inf['daily_excnet_simp']).plot(alpha=1, ax=a, lw=1)
    date_utils.chg_idx2datestr(result_analyse_effect['longshort_inf']['net_l_s_ls']['多空']).plot(alpha=1, ax=a, lw=1, color='cornflowerblue')
    plt.legend(loc='center left', bbox_to_anchor=(1, 0.5), fontsize='x-small', labels=['g' + str(i) for i in
                                                            group_ret_inf['daily_excnet_simp'].columns.to_list()] + ['long_short'])
    a.set_title('Group Excess Return (Simp)')
    plt.close(f)
    ws.pictures.add(f, name='Group Excess Return', update=True, left=ws.range('A55').left, top=ws.range('A55').top)

def save_score_by_size_ind(result_factor_analyse):
    global wb, current_ws
    result_score_by_size_ind = result_factor_analyse['score_by_size_ind']
    print('输出市值行业分层打分...')
    ws = wb.sheets.add('市值行业分层打分', after=current_ws)
    ws.range('A1').value = '分组收益_不对冲'
    ws.range('B1').value = result_score_by_size_ind['eva_longshort_simp']

    #
    formatrange(ws, 'A1', "", 'Center', if_Bold=True)
    formatrange(ws, 'C:H', "", 'Center')
    formatrange(ws, 'C2:H8', "0.00%", 'Center')
    formatrange(ws, 'C4:H4', "0.00", 'Center')
    formatrange(ws, 'C6:H8', "0.00", 'Center')


    get_rs_cols(ws)
    ws.range(get_column_letter(c_count + 2) + '1').value = '分组绝对净值'
    #和多空净值合并
    ws.range(get_column_letter(c_count + 3) + '1').value = date_utils.chg_idx2datestr(result_score_by_size_ind['daily_net_simp'])
    ws.range(get_column_letter(c_count + 4) + '1').value = ['组' + str(i) for i in
                                                            result_score_by_size_ind['daily_net_simp'].columns.to_list()]
    formatrange(ws, get_column_letter(c_count + 4) + ':' + get_column_letter(c_count + 9), "0.00", 'Center')

    get_rs_cols(ws)
    ws.range(get_column_letter(c_count + 2) + '1').value = '分组每期收益'
    ws.range(get_column_letter(c_count + 3) + '1').value = date_utils.chg_idx2datestr(result_score_by_size_ind['group_ret'])
    ws.range(get_column_letter(c_count + 4) + '1').value = ['组' + str(i) for i in
                                                        result_score_by_size_ind['group_ret'].columns.to_list()]
    formatrange(ws, get_column_letter(c_count + 4) + ':' + get_column_letter(c_count + 8), "0.00%", 'Center')
    #画图：
    #1.分组收益
    sns.set()
    f, a = plt.subplots(1, 1, figsize=(4, 2))
    g_num = result_score_by_size_ind['eva_longshort_simp'].shape[1] - 1
    plt.bar(np.arange(g_num), result_score_by_size_ind['eva_longshort_simp'].loc['AnnualRt'][:-1],
            tick_label=list(range(1, g_num+1)))
    a.set_title('Group Return')
    plt.close(f)
    ws.pictures.add(f, name='group return', update=True, left=ws.range('A25').left, top=ws.range('A14').top)


    #2.每组净值
    sns.set()
    sns.set_palette(sns.cubehelix_palette(result_score_by_size_ind['daily_net_simp'].shape[1]-1))
    fig, ax1 = plt.subplots(figsize=(8, 5), dpi=100)
    ax2 = ax1.twinx()
    date_utils.chg_idx2datestr(result_score_by_size_ind['daily_net_simp'].iloc[:,:-1]).plot(alpha=1, ax=ax1, lw=1)
    plt.legend(loc='upper right', bbox_to_anchor=(0, 0.5), fontsize='x-small', labels=['g' + str(i) for i in
                                                            result_score_by_size_ind['daily_net_simp'].columns[:-1].to_list()])

    date_utils.chg_idx2datestr(result_score_by_size_ind['daily_net_simp'].iloc[:, -1]).plot(alpha=1, ax=ax2, lw=1, color='cornflowerblue')
    plt.legend(loc='upper right', bbox_to_anchor=(1, 0.5), fontsize='x-small', labels= ['long_short(Right axis)'])

    ax1.set_title('Group Return (Simp)-Score by Size and Industry')
    plt.close(fig)
    ws.pictures.add(fig, name='Score by Size and Industry', update=True, left=ws.range('A28').left, top=ws.range('A28').top)


    current_ws = '市值行业分层打分'

def save_corr(result_factor_analyse):
    global wb, current_ws
    corr_all = result_factor_analyse['corr_all']
    ws = wb.sheets.add('与风险因子相关性', after=current_ws)
    ws.range('A1').value = '与风险因子相关性'
    ws.range('B1').value = corr_all['mean']
    get_rs_cols(ws)
    ws.range('A{}'.format(r_count + 2)).value = '稳定系数'
    ws.range('B{}'.format(r_count + 2)).value = corr_all['stability']

    formatrange(ws, 'A:A', "", 'Center', if_Bold=True)
    formatrange(ws, 'C:' + get_column_letter(c_count), "0.00", 'Center')


    sns.set(style="white")
    # Generate a mask for the upper triangle
    mask = np.zeros_like(corr_all['mean'], dtype=np.bool)
    mask[np.triu_indices_from(mask)] = True
    # Set up the matplotlib figure
    f, ax = plt.subplots(figsize=(9, 7))
    # Generate a custom diverging colormap
    cmap = sns.diverging_palette(220, 10, as_cmap=True)
    # Draw the heatmap with the mask and correct aspect ratio
    sns.heatmap(corr_all['mean'], mask=mask, cmap=cmap, vmax=.3, center=0,
                square=True, linewidths=.5, cbar_kws={"shrink": .5}, annot=True, fmt='.2f')
    ax.set_title('Correlation with Risk Factors')
    plt.ylabel('')
    plt.close(f)
    ws.pictures.add(f, name='Correlation with Risk Factors', update=True, left=ws.range('O1').left,
                    top=ws.range('O1').top)

    #
    sns.set(style="white")
    # Generate a mask for the upper triangle
    mask = np.zeros_like(corr_all['stability'], dtype=np.bool)
    mask[np.triu_indices_from(mask)] = True
    # Set up the matplotlib figure
    f, ax = plt.subplots(figsize=(9, 7))
    # Generate a custom diverging colormap
    cmap = sns.diverging_palette(220, 10, as_cmap=True)
    # Draw the heatmap with the mask and correct aspect ratio
    sns.heatmap(corr_all['stability'], mask=mask, cmap=cmap, vmax=.3, center=0,
                square=True, linewidths=.5, cbar_kws={"shrink": .5}, annot=True, fmt='.2f')
    ax.set_title('Correlation Stability with Risk Factors')
    plt.ylabel('')
    plt.close(f)
    ws.pictures.add(f, name='Correlation Stability with Risk Factors_2', update=True, left=ws.range('AF1').left,
                    top=ws.range('AF1').top)

    current_ws = '与风险因子相关性'

def save_testresult(result_factor_analyse):
    global r_count, c_count, wb, current_ws

    app = xw.App()
    wb = app.books.add() #新建workbook

    if_run = {'analyse_effect': False}


    print('输出配置信息...')
    save_config(result_factor_analyse)

    if ('analyse_effect' in result_factor_analyse.keys()):
        if bool(result_factor_analyse['analyse_effect']):
            if_run['analyse_effect'] = True
            result_analyse_effect = result_factor_analyse['analyse_effect']
            print("输出IC...")
            save_ic(result_analyse_effect )

            print('输出因子收益...')
            save_fac_ret(result_analyse_effect )

            print('输出多空信息...')
            save_longshort_inf(result_analyse_effect )

            print('输出分组信息...')
            save_group_inf(result_analyse_effect )


    if ('score_by_size_ind' in result_factor_analyse.keys()):
        if (bool(result_factor_analyse['score_by_size_ind'])):
            print('输出市值行业分层打分...')
            save_score_by_size_ind(result_factor_analyse)

    if ('corr_all' in result_factor_analyse.keys()):
        print('输出与风险因子相关性')
        save_corr(result_factor_analyse)

    if ('analyse_effect_bysample' in result_factor_analyse.keys()):
        if (bool(result_factor_analyse['analyse_effect_bysample'])):
            if if_run['analyse_effect']:
                print('输出分样本信息')
                sample_all = list(result_factor_analyse['analyse_effect_bysample'].keys())
                for sample_i in sample_all:
                    result_sample = result_factor_analyse['analyse_effect_bysample'][sample_i]
                    print(sample_i, "输出IC...")
                    print(current_ws)
                    save_ic(result_sample, sample_i)

                    print(sample_i, "输出因子收益...")
                    print(current_ws)
                    save_fac_ret(result_sample, sample_i)

                    print(sample_i, "输出多空信息...")
                    print(current_ws)
                    save_longshort_inf(result_sample, sample_i)

                    print(sample_i, "输出分组信息...")
                    print(current_ws)
                    save_group_inf(result_sample, sample_i)


    print('保存结果...')
    excelname = '单因子测试_{}_{}_{}.xlsx'.format(result_factor_analyse['factor_setting']['name'],
                                                  result_factor_analyse['prepro_setting']['sample_index'],
                                                     result_factor_analyse['execute_setting']['EffectAnalysis_hedge'])
    wb.save(result_factor_analyse['execute_setting']['save_filedir'] + excelname)
    # wb.close()
    app.quit()

if __name__ == '__main__':
    from factor_config import (
            factor_setting,
            prepro_setting,
            execute_setting)
    result_factor_analyse = factor_performance.analyse_performance(factor_setting, prepro_setting, execute_setting)
    save_testresult(result_factor_analyse)

